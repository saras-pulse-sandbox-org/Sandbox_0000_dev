import base64
import json
import logging
import os
import tempfile
from datetime import datetime
from pathlib import Path

import yaml
from google.cloud import pubsub_v1
from google.oauth2 import service_account
from openpyxl import Workbook

PROJECT_ID = os.getenv("BQ_PROJECT_ID", "solutionsdw")

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger("dag_alerts")
profile_filepath = Path(__file__).parent.parent / "profiles.yml"
profiles_hash = yaml.safe_load(open(profile_filepath, "r"))

PROJECT_NAME = [*profiles_hash][0]
SLACK_CONN_ID = "slack_saras"
SLACK_TOPIC = "pulse-send-slack"
ALERTS_CHANNEL = "C06HT2JK96X" if os.getenv("ENVIRONMENT") == "prod" else "C065MG2L63U"
DBT_TEST_ALERTS_CHANNEL = "C07SY7G2B47" if os.getenv("ENVIRONMENT") == "prod" else "C065MG2L63U"
CLICKUP_TOPIC = os.getenv("CLICKUP_TOPIC", "pulse-create-ticket")

dags_dir = f"/home/airflow/gcs/dags"
client_dir = f"{dags_dir}/{PROJECT_NAME}" if os.getenv("ENVIRONMENT") == "prod" else dags_dir
credentials_path = f"{client_dir}/keyfile.json"


class DagRunFailException(Exception):
    """
    Exception raised when one or more tasks in a DAG run fail.
    """

    ...


def create_clickup_alert(
    task: str,
    dag_id: str,
    exec_date: datetime,
    log_url: str,
    logs: str,
):
    """
    Sends message to clickup topic upon failure
    Args:
        task (str): The task ID.
        dag_id (str): The DAG ID.
        log_url (str): The log URL.
    Returns:
        None
    """
    # Create credentials using the service account file and
    # initialize Pub/Sub client with the specified credentials
    credentials = service_account.Credentials.from_service_account_file(credentials_path)
    pubsub_publisher = pubsub_v1.PublisherClient(credentials=credentials)
    topic_path = pubsub_publisher.topic_path(PROJECT_ID, CLICKUP_TOPIC)

    client_name, client_id, _ = dag_id.split("_")
    exec_time = int(exec_date.timestamp())
    exec_time_fallback = exec_date.strftime(r"%Y-%m-%d %I:%M:%S %p")

    list_id = "901601915463"
    title = f"DAG failure: {dag_id}.{task}"
    description = (
        f"Client Name: {client_name}\n"
        f"Client ID: {client_id}\n"
        f"DAG ID: {dag_id}\n"
        f"Task ID: {task}\n"
        f"Log URL: {log_url}"
    )
    tags = ["automation", "dag-failure", client_name, client_id]
    priority = 2

    logger.info("Sending clickup alert")
    message = {
        "api_args": {
            "name": title,
            "content": description,
            "tags": tags,
            "priority": priority,
            "due_date": int(datetime.now().timestamp() * 1000),
            "start_date": int(datetime.now().timestamp() * 1000),
            "assignees": [67469374, 61256138, 55259537],
        },
        "list_id": list_id,
        "log_url": log_url,
        "dag_id": dag_id,
        "task": task,
        "logs": None,
        "exec_time": exec_time,
        "exec_time_fallback": exec_time_fallback,
    }
    logger.info("Message: %s", message)
    future = pubsub_publisher.publish(topic_path, json.dumps(message).encode("utf-8"))
    future.result()
    logger.info("Published message to %s.", CLICKUP_TOPIC)


def on_task_fail(context):
    """
    This is a callback function that is triggered when a task fails in a DAG.
    Currently this function handles sending Slack alerts and creating Jira tickets for failed tasks.

    Args:
        context (dict): The Airflow context.
    Returns:
        None
    """
    logger.info("Executing On Failure Callback")
    logger.info("Context: %s", context)
    is_manual_trigger = context["run_id"].startswith("manual")

    # Ignore if the task is manually triggered
    if is_manual_trigger:
        logging.info("Manual trigger. Skipping On Failure Callback")
        return

    # Get the exception
    exception = context.get("exception")

    # DagRunFailException is raised when one or more upstream tasks fail
    # So skip the alert if this exception is raised since the upstream
    # tasks alerts would have been sent
    if exception and isinstance(exception, DagRunFailException):
        logging.info("Alerts for the failed upstream tasks have been sent already. Skipping On Failure Callback.")
        return

    task_instance = context.get("task_instance")
    dag_id = task_instance.dag_id
    log_url = task_instance.log_url
    task_id = task_instance.task_id
    exec_date = context.get("execution_date")
    cli_logs = context.get("cli_logs")

    logger.info("Task ID: %s", task_id)
    logger.info("Log URL: %s", log_url)
    logger.info("Execution Date: %s", exec_date)
    logger.info("Exception: %s", exception)
    logger.info("CLI Logs: %s", cli_logs)
    try:
        task = task_id.split(".", maxsplit=1)[0]
    except ValueError:
        task = task_id

    # Create Clickup ticket and Send slack alert
    create_clickup_alert(task, dag_id, exec_date, log_url, cli_logs)


def parse_dbt_test_results(run_results_path, manifest_path):
    """
    Parse dbt test results from the run_results.json and manifest.json files.

    Args:
        run_results_path (str): Path to the run_results.json file.
        manifest_path (str): Path to the manifest.json file.

    Returns:
        dict: A dictionary containing parsed test results.
    """
    with open(run_results_path, "r") as f:
        run_results = json.load(f)

    with open(manifest_path, "r") as f:
        manifest = json.load(f)

    total_tests = len(run_results["results"])
    passed_count = sum(1 for result in run_results["results"] if result["status"] == "pass")
    failed_count = sum(1 for result in run_results["results"] if result["status"] == "fail")
    warn_count = sum(1 for result in run_results["results"] if result["status"] == "warn")

    unsuccessful_tests = []
    for result in run_results["results"]:
        if result["status"] in ["fail", "warn"]:
            unique_id = result["unique_id"]
            node = manifest["nodes"].get(unique_id, {})

            model_name = node.get("depends_on", {}).get("nodes", [None])[0]
            if model_name:
                model_name = model_name.split(".")[-1]
            else:
                model_name = "Unknown"

            # Use relation_name directly for the test failure query
            test_failure_query = node.get("relation_name", "No relation name available")

            unsuccessful_tests.append(
                {
                    "test_name": unique_id.split(".")[-2],
                    "model_name": model_name,
                    "column_name": node.get("column_name"),
                    "test_type": node.get("test_metadata", {}).get("name", "Unknown"),
                    "file_path": node.get("original_file_path", "Unknown"),
                    "failures": result.get("failures", 0),
                    "message": result.get("message", "No message available"),
                    "severity": node.get("config", {}).get("severity", "ERROR"),
                    "tags": node.get("tags", []),
                    "status": result["status"],
                    "test_failure_query": f"select * from {test_failure_query}",
                    "compiled_code": node.get("compiled_code", "No compiled code available"),
                }
            )

    return {
        "total_tests": total_tests,
        "passed_count": passed_count,
        "failed_count": failed_count,
        "warn_count": warn_count,
        "unsuccessful_tests": unsuccessful_tests,
    }


def create_excel_report(test_results):
    """
    Create an Excel file with detailed test results.

    Args:
        test_results (dict): Parsed dbt test results.

    Returns:
        bytes: Excel file content as bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "DBT Test Results"

    # Add headers
    headers = ["Model", "Column", "Test Type", "Failures", "Status", "Message", "Test Failure Query", "Compiled Code"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Add data
    for row, test in enumerate(test_results["unsuccessful_tests"], start=2):
        ws.cell(row=row, column=1, value=test["model_name"])
        ws.cell(row=row, column=2, value=test["column_name"])
        ws.cell(row=row, column=3, value=test["test_type"])
        ws.cell(row=row, column=4, value=test["failures"])
        ws.cell(row=row, column=5, value=test["status"].upper())
        ws.cell(row=row, column=6, value=test["message"])
        ws.cell(row=row, column=7, value=test["test_failure_query"])
        ws.cell(row=row, column=8, value=test["compiled_code"])

    # Save the file to a bytes buffer
    with tempfile.NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        return tmp.read()


def send_slack_alert(context, test_results):
    """
    Send a Slack alert with summarized dbt test results and attach an Excel file with details.

    Args:
        context (dict): The Airflow context.
        test_results (dict): Parsed dbt test results.
    """
    task_instance = context.get("task_instance")
    dag_id = task_instance.dag_id
    log_url = task_instance.log_url
    exec_date = context.get("execution_date")

    client_name, client_id = dag_id.rsplit("_", 2)[:2]

    # Create Excel report
    excel_content = create_excel_report(test_results)
    encoded_file = base64.b64encode(excel_content).decode("utf-8")

    message = {
        "channel_id": DBT_TEST_ALERTS_CHANNEL,
        "username": "DBT Test Results",
        "icon_emoji": ":dbt:",
        "blocks": [
            {"type": "context", "elements": [{"type": "mrkdwn", "text": f"{client_name} (ID: {client_id})"}]},
            {"type": "divider"},
            {
                "type": "section",
                "fields": [
                    {"type": "mrkdwn", "text": f"*:bar_chart: Total Tests:* {test_results['total_tests']}"},
                    {"type": "mrkdwn", "text": f"*:white_check_mark: Passed:* {test_results['passed_count']}"},
                    {"type": "mrkdwn", "text": f"*:x: Failed:* {test_results['failed_count']}"},
                    {"type": "mrkdwn", "text": f"*:warning: Warning:* {test_results['warn_count']}"},
                ],
            },
            {
                "type": "actions",
                "elements": [{"type": "button", "text": {"type": "plain_text", "text": "View Logs"}, "url": log_url}],
            },
        ],
        "attachments": [
            {
                "filename": "dbt_test_results.xlsx",
                "content": encoded_file,
                "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
        ],
    }

    # Use the existing Pub/Sub setup to send the Slack message
    logger.info("Sending Slack alert for DBT test results")
    logger.info("Credentials path: %s", credentials_path)
    credentials = service_account.Credentials.from_service_account_file(credentials_path)
    pubsub_publisher = pubsub_v1.PublisherClient(credentials=credentials)
    topic_path = pubsub_publisher.topic_path(PROJECT_ID, SLACK_TOPIC)
    logger.info("Project ID: %s", PROJECT_ID)
    logger.info("SLACK_TOPIC: %s", SLACK_TOPIC)
    logger.info("Topic path: %s", topic_path)

    # Encode the entire message as JSON and publish it
    encoded_message = json.dumps(message).encode("utf-8")
    future = pubsub_publisher.publish(topic_path, encoded_message)
    future.result()
    logger.info("Published Slack alert for DBT test results.")


def on_dbt_test_completion(context):
    """
    Callback function for dbt test task completion.

    Args:
        context (dict): The Airflow context.
    """
    logger.info("Executing on_dbt_test_completion callback")
    is_manual_trigger = context["run_id"].startswith("manual")

    # Ignore if the task is manually triggered
    if is_manual_trigger:
        logger.info("Manual trigger. Skipping On Failure Callback")
        return

    run_results_path = Path(client_dir) / "target" / "run_results.json"
    manifest_path = Path(client_dir) / "target" / "manifest.json"

    if run_results_path.exists() and manifest_path.exists():
        logger.info("run_results.json and manifest.json found in {client_dir}/target/")
        test_results = parse_dbt_test_results(str(run_results_path), str(manifest_path))

        # Only send Slack alert if there are failed tests
        if test_results["failed_count"] > 0:
            logger.info("Sending Slack alert for failed tests")
            send_slack_alert(context, test_results)
        else:
            logger.info("All dbt tests passed. No Slack alert sent.")
    else:
        logger.error(f"run_results.json or manifest.json not found in {client_dir}/target/")
